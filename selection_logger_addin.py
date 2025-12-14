"""ArcGIS Pro Python Add-in button to log selected feature counts to Excel.

This script is designed for ArcGIS Pro's Python add-in framework. The button
searches for a target layer in the active map, counts the currently selected
features, and appends the count to an Excel workbook. If the workbook does not
exist, it is created with a header row.

Configuration notes:
- Update TARGET_LAYER_NAME to match the layer you want to monitor.
- Update EXCEL_LOG_PATH to the desired output Excel (.xlsx) path.
- Register this script in the add-in's config.xml (see the accompanying file).
"""

import os
from datetime import datetime

import arcpy
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# User configuration
# ---------------------------------------------------------------------------
# Name of the shapefile/feature layer in the active map that the button should
# inspect. Use the name as it appears in the Contents pane.
TARGET_LAYER_NAME = "MyShapefileLayer"

# Location where the Excel log should be written. The folder will be created
# automatically if it does not already exist.
EXCEL_LOG_PATH = r"C:\\GIS\\SelectionLogs\\selection_log.xlsx"


class SelectionLoggerButton(object):
    """Button implementation for logging selection counts.

    The class name must match the "class" attribute in config.xml. ArcGIS Pro
    loads this class and calls onClick when the user presses the button.
    """

    def __init__(self):
        # Button state flags required by the add-in framework
        self.enabled = True
        self.checked = False

    def onClick(self):
        """Entry point executed when the button is clicked."""
        try:
            # Access the current ArcGIS Pro project and the active map view.
            aprx = arcpy.mp.ArcGISProject("CURRENT")
            active_map = aprx.activeMap
            if active_map is None:
                raise RuntimeError("No active map found in the current project.")

            # Locate the target layer by name.
            target_layer = _find_layer_by_name(active_map, TARGET_LAYER_NAME)
            if target_layer is None:
                raise RuntimeError(
                    f"Layer '{TARGET_LAYER_NAME}' was not found in the active map."
                )

            # Count ONLY the currently selected features.
            #
            # IMPORTANT:
            #   arcpy.management.GetCount(layer) often returns the *total* features
            #   in the layer (ignoring selection) when working with ArcGIS Pro
            #   layer objects.
            #
            # The reliable way is to read the selection set from Describe(...).FIDSet.
            # FIDSet is a semicolon-separated list of selected ObjectIDs.
            selected_count = _count_selected_features(target_layer)

            # Append the count to Excel (creating the file if needed).
            log_path = _log_selection_to_excel(
                layer_name=target_layer.name,
                selection_count=selected_count,
                excel_path=EXCEL_LOG_PATH,
            )

            arcpy.AddMessage(
                f"Selection count logged: {selected_count} features (file: {log_path})."
            )
        except PermissionError:
            # Common when the Excel workbook is open in another application.
            arcpy.AddError(
                "The Excel log is currently locked. Close the workbook and try again."
            )
        except Exception as exc:  # Broad catch to surface any issues to the user.
            arcpy.AddError(f"Selection logging failed: {exc}")


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _find_layer_by_name(map_obj, layer_name):
    """Return the first layer in the map whose name matches layer_name."""
    for layer in map_obj.listLayers():
        if layer.name == layer_name:
            return layer
    return None


def _count_selected_features(layer) -> int:
    """Return the number of selected features for a feature layer.

    Uses Describe(layer).FIDSet which is a semicolon-separated string of selected
    ObjectIDs. If nothing is selected, FIDSet is an empty string.

    This is more reliable than GetCount(layer) for selection counting in Pro.
    """
    desc = arcpy.Describe(layer)
    fidset = getattr(desc, "FIDSet", "")

    if not fidset:
        return 0

    # FIDSet is usually ';' separated, but we defensively normalize commas too.
    parts = [p.strip() for p in fidset.replace(",", ";").split(";") if p.strip()]
    return len(parts)


def _ensure_directory(path):
    """Create the parent directory for the Excel path if it does not exist."""
    parent_dir = os.path.dirname(path)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)


def _log_selection_to_excel(layer_name, selection_count, excel_path):
    """Append a selection count record to an Excel workbook.

    Args:
        layer_name (str): Name of the layer used for the selection.
        selection_count (int): Number of selected features at the time of logging.
        excel_path (str): Full path to the Excel workbook.

    Returns:
        str: The Excel path used for logging (useful for messaging).
    """

    _ensure_directory(excel_path)

    # Initialize or open the workbook.
    #
    # IMPORTANT:
    #   Do NOT rely on workbook.active for existing files, because the active
    #   sheet can change (e.g., if a user saved the workbook with a different
    #   sheet selected). That can cause "event numbers" to restart or headers
    #   to be missing.
    sheet_name = "Selection Log"
    if not os.path.exists(excel_path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
    else:
        workbook = load_workbook(excel_path)
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # Ensure headers exist (first row)
    if worksheet.max_row < 1 or (worksheet.cell(1, 1).value is None and worksheet.cell(1, 2).value is None):
        worksheet.append(["Event", "Count", "Layer", "Timestamp"])

    # Determine the next event number (robustly).
    # We parse the last "Selection Event N" value so numbers never repeat even
    # if rows are blank or the sheet was edited.
    next_event = _next_event_number(worksheet)
    event_label = f"Selection Event {next_event}"

    # Keep count as a NUMBER in Excel (better for formulas).
    count_value = int(selection_count)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    worksheet.append([event_label, count_value, layer_name, timestamp])
    workbook.save(excel_path)

    return excel_path


def _next_event_number(ws) -> int:
    """Return the next sequential Selection Event number.

    Looks from bottom to top in column A for values like "Selection Event N".
    If none found, returns 1.
    """
    for row in range(ws.max_row, 1, -1):  # skip header row
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and "Selection Event" in val:
            # Try to parse the trailing integer.
            try:
                n = int(val.strip().split()[-1])
                return n + 1
            except Exception:
                continue
    return 1


# End of file
